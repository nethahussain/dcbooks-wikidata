#!/usr/bin/env python3
"""
Link DC Books authors to Wikidata items.
=========================================
Looks up each unique author in the dataset against Wikidata's search API,
finds matching human items, and adds a P50_QID column to the Excel file.

Run locally (requires network access to wikidata.org):
  python3 link_authors_wikidata.py
  python3 link_authors_wikidata.py --input data/dcbooks_wikidata.xlsx
"""

import requests, json, time, re, argparse, os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

CACHE_FILE = "author_qid_cache.json"
HEADERS = {"User-Agent": "DCBooksWikidataBot/1.0 (nethahussain@gmail.com)"}

# Load/save cache
def load_cache():
    try:
        with open(CACHE_FILE) as f:
            return json.load(f)
    except:
        return {}

def save_cache(cache):
    with open(CACHE_FILE, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=1)


def search_wikidata(name, cache):
    """Search Wikidata for a person matching the author name."""
    if name in cache:
        return cache[name]

    time.sleep(0.3)  # rate limit
    try:
        resp = requests.get("https://www.wikidata.org/w/api.php",
            params={
                "action": "wbsearchentities",
                "search": name,
                "language": "en",
                "type": "item",
                "limit": 5,
                "format": "json",
            },
            headers=HEADERS, timeout=15)
        data = resp.json()

        for result in data.get("search", []):
            qid = result["id"]
            desc = result.get("description", "").lower()
            label = result.get("label", "")

            # Match criteria: must be a person (writer, poet, etc.)
            person_kw = [
                "writer", "author", "poet", "novelist", "journalist",
                "activist", "politician", "actor", "actress", "director",
                "singer", "artist", "philosopher", "scholar", "editor",
                "translator", "critic", "historian", "scientist",
                "educator", "teacher", "professor", "leader",
                "playwright", "screenwriter", "lyricist", "essayist",
                "biographer", "humorist", "satirist", "cartoonist",
                "indian", "malayalam", "kerala", "bengali", "hindi",
                "tamil", "telugu", "marathi", "gujarati",
                "american", "british", "french", "german", "russian",
                "japanese", "chinese", "spanish", "italian",
                "nobel", "booker", "pulitzer", "sahitya akademi",
                "born in", "born on",
            ]
            if any(kw in desc for kw in person_kw):
                cache[name] = qid
                return qid

            # Also accept if label matches exactly and it's a human
            if label.lower() == name.lower() and desc and "disambiguation" not in desc:
                # Verify it's a human via the API
                verify = verify_human(qid)
                if verify:
                    cache[name] = qid
                    return qid

        cache[name] = ""
        return ""
    except Exception as e:
        print(f"  Error searching '{name}': {e}")
        return ""


def verify_human(qid):
    """Quick check if a QID is a human (P31=Q5)."""
    try:
        resp = requests.get("https://www.wikidata.org/w/api.php",
            params={
                "action": "wbgetclaims",
                "entity": qid,
                "property": "P31",
                "format": "json",
            },
            headers=HEADERS, timeout=10)
        data = resp.json()
        claims = data.get("claims", {}).get("P31", [])
        for claim in claims:
            target = claim.get("mainsnak", {}).get("datavalue", {}).get("value", {}).get("id", "")
            if target == "Q5":  # human
                return True
    except:
        pass
    return False


def main():
    parser = argparse.ArgumentParser(description="Link DC Books authors to Wikidata QIDs")
    parser.add_argument("--input", default="data/dcbooks_wikidata.xlsx",
                        help="Input Excel file")
    args = parser.parse_args()

    if not os.path.exists(args.input):
        print(f"File not found: {args.input}")
        print("Run from the repo root directory, or specify --input path")
        return

    cache = load_cache()
    print(f"Loaded cache: {len(cache)} entries ({sum(1 for v in cache.values() if v)} with QIDs)")

    # Load workbook
    print(f"Loading {args.input}...")
    wb = load_workbook(args.input)
    ws = wb["Wikidata Upload"]
    ws2 = wb["Raw Data"]
    total = ws.max_row - 1

    # Find/create columns
    headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    col_author = headers["P50 (author)"]
    col_qid = headers.get("P50_QID (author Wikidata ID)")
    if not col_qid:
        col_qid = ws.max_column + 1
        cell = ws.cell(1, col_qid, value="P50_QID (author Wikidata ID)")
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        cell.fill = PatternFill("solid", fgColor="2B579A")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=Side("thin"), right=Side("thin"),
                            top=Side("thin"), bottom=Side("thin"))

    # Get unique authors
    authors = set()
    for r in range(2, ws.max_row + 1):
        a = ws.cell(r, col_author).value
        if a and a.strip():
            authors.add(a.strip())

    uncached = [a for a in authors if a not in cache]
    print(f"Unique authors: {len(authors)}, uncached: {len(uncached)}")

    # Look up each author
    found = 0
    for i, name in enumerate(sorted(authors)):
        qid = search_wikidata(name, cache)
        if qid:
            found += 1

        if (i + 1) % 50 == 0:
            save_cache(cache)
            cached_found = sum(1 for a in authors if cache.get(a))
            print(f"  {i+1}/{len(authors)} searched, {cached_found} found so far")

    save_cache(cache)

    # Write QIDs to Excel
    rows_filled = 0
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, col_author).value
        if name:
            qid = cache.get(name.strip(), "")
            if qid:
                ws.cell(r, col_qid).value = qid
                rows_filled += 1

    wb.save(args.input)

    # Summary
    matched = sum(1 for a in authors if cache.get(a))
    print(f"\n{'='*50}")
    print(f"Authors with Wikidata QID: {matched}/{len(authors)} unique")
    print(f"Book rows with author QID: {rows_filled}/{total}")
    print(f"Saved to {args.input}")
    print(f"Cache saved to {CACHE_FILE}")

    # Show matches
    print(f"\nMatched authors ({matched}):")
    for name in sorted(authors):
        qid = cache.get(name, "")
        if qid:
            print(f"  {name:40s} → {qid}")


if __name__ == "__main__":
    main()
