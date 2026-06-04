#!/usr/bin/env python3
"""
Scrape keralabookstore.com book metadata -> Wikidata-ready spreadsheet.

USAGE (run LOCALLY - the site rate-limits datacenter/cloud IPs with HTTP 429):
    pip install isbnlib openpyxl
    python3 scrape_keralabookstore_full.py            # scrape ALL books
    python3 scrape_keralabookstore_full.py 0 500      # scrape first 500 (start count)

Outputs:
    keralabookstore_books.jsonl   (raw scrape, resumable)
    keralabookstore_wikidata.xlsx (upload-ready spreadsheet)
"""
import re, html, json, time, sys, os, urllib.request
import isbnlib
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

UA="Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120 Safari/537.36"
SITEMAP="https://keralabookstore.com/sitemap.xml"
JSONL="keralabookstore_books.jsonl"
XLSX="keralabookstore_wikidata.xlsx"
DELAY=1.0          # politeness delay (seconds) — raise if you see 429s
LANG={"Malayalam":"Q36236","English":"Q1860","Hindi":"Q1568","Tamil":"Q5885","Sanskrit":"Q11059","Arabic":"Q13955"}
FMT={"Paperback":"Q193934","Hardcover":"Q193955","Hardback":"Q193955"}

def fetch(url, tries=4):
    for k in range(tries):
        try:
            r=urllib.request.Request(url, headers={"User-Agent":UA})
            with urllib.request.urlopen(r, timeout=30) as resp:
                return resp.read().decode("utf-8","ignore")
        except urllib.error.HTTPError as e:
            if e.code==429: time.sleep(30*(k+1))   # backoff hard on rate limit
            else: time.sleep(3*(k+1))
        except Exception:
            time.sleep(3*(k+1))
    return None

def clean(s): return html.unescape(re.sub(r'\s+',' ',s)).strip()

def parse(h, url):
    def first_name():
        m=re.search(r'itemprop="name"[^>]*>([^<]+)<', h); return clean(m.group(1)) if m else ""
    def after(prop):
        m=re.search(r'itemprop="%s".*?itemprop="name"[^>]*>([^<]+)<'%prop, h, re.S); return clean(m.group(1)) if m else ""
    def ip(prop):
        m=re.search(r'itemprop="%s"[^>]*content="([^"]*)"'%prop, h)
        if m: return clean(m.group(1))
        m=re.search(r'itemprop="%s"[^>]*>([^<]*)<'%prop, h); return clean(m.group(1)) if m else ""
    sm=re.search(r'/book/([^/]+)/(\d+)', url); slug=sm.group(1) if sm else ""; bid=sm.group(2) if sm else ""
    isbn=ip("isbn")
    if not re.search(r'\d{10}', isbn):
        m=re.search(r'\b(97[89]\d{10})\b', h); isbn=clean(m.group(1)) if m else ""
    isbn=re.sub(r'[^0-9Xx]','',isbn)
    ed=ip("bookEdition"); ym=re.search(r'(\d{4})', ed)
    return {"id":bid,"url":url,"title_ml":first_name(),"title_en":clean(slug.replace("-"," ")).title(),
            "author":after("author"),"publisher":after("publisher"),"language":after("inLanguage"),
            "isbn":isbn,"pages":re.sub(r'\D','',ip("numberOfPages")),"edition_year":ym.group(1) if ym else "",
            "format":ip("bookFormat").split("/")[-1]}

def isbn13h(s):
    s=re.sub(r'[^0-9Xx]','',s or "")
    if isbnlib.is_isbn13(s): return isbnlib.mask(s)
    if isbnlib.is_isbn10(s):
        t=isbnlib.to_isbn13(s); return isbnlib.mask(t) if t else ""
    return ""
def valid(s): s=re.sub(r'[^0-9Xx]','',s or ""); return isbnlib.is_isbn13(s) or isbnlib.is_isbn10(s)

def build_xlsx(recs):
    wb=Workbook(); ws=wb.active; ws.title="KeralaBookStore Upload"
    cols=["id","url","Len (title en)","Lml (title ml)","P2093 author","publisher","language",
          "P407 lang QID","P212 ISBN-13 (hyphenated)","ISBN valid?","P1104 pages","P577 date",
          "P437 format QID","format","S854 source"]
    ws.append(cols)
    for c in ws[1]:
        c.font=Font(bold=True,color="FFFFFF"); c.fill=PatternFill("solid",start_color="305496")
        c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    ws.freeze_panes="A2"
    for r in recs:
        v=valid(r.get("isbn","")); yr=r.get("edition_year","")
        date=f"+{yr}-01-01T00:00:00Z/9" if re.match(r'^\d{4}$',yr or "") else ""
        ws.append([r.get("id",""),r.get("url",""),r.get("title_en",""),r.get("title_ml",""),r.get("author",""),
                   r.get("publisher",""),r.get("language",""),LANG.get(r.get("language",""),""),
                   isbn13h(r.get("isbn","")),"yes" if v else "NO",r.get("pages",""),date,
                   FMT.get(r.get("format",""),""),r.get("format",""),r.get("url","")])
    for i,w in enumerate([9,42,26,26,22,24,11,11,22,9,9,22,12,12,42],1):
        ws.column_dimensions[get_column_letter(i)].width=w
    wb.save(XLSX)

def main():
    start=int(sys.argv[1]) if len(sys.argv)>1 else 0
    count=int(sys.argv[2]) if len(sys.argv)>2 else 10**9
    print("Fetching sitemap…")
    sm=fetch(SITEMAP)
    urls=re.findall(r'<loc>(https://keralabookstore\.com/book/[^<]+)</loc>', sm)
    print(f"{len(urls)} book URLs; scraping [{start}:{start+count}] with {DELAY}s delay")
    done={json.loads(l)["url"] for l in open(JSONL,encoding="utf-8")} if os.path.exists(JSONL) else set()
    target=urls[start:start+count]
    with open(JSONL,"a",encoding="utf-8") as f:
        for i,u in enumerate(target):
            if u in done: continue
            h=fetch(u)
            if h: f.write(json.dumps(parse(h,u),ensure_ascii=False)+"\n"); f.flush()
            if (i+1)%50==0: print(f"  {i+1}/{len(target)}", flush=True)
            time.sleep(DELAY)
    recs=[json.loads(l) for l in open(JSONL,encoding="utf-8") if l.strip()]
    build_xlsx(recs)
    print(f"Done. {len(recs)} books -> {XLSX}")

if __name__=="__main__": main()
