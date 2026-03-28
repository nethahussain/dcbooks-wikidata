#!/usr/bin/env python3
"""
DC Books Full Metadata Scraper → Wikidata-ready Excel
=====================================================
1. Fetches all book URLs from dcbookstore.com/sitemap.xml (~3800+ books)
2. Solves Sucuri WAF challenge via Node.js
3. Scrapes each book page for full metadata:
   title, author, ISBN, category, binding, pub date, publisher, edition,
   pages, language, summary (often in Malayalam)
4. Outputs formatted Excel ready for Wikidata QuickStatements upload

Usage:
  python dcbooks_scraper.py                         # scrape ALL books (several hours)
  python dcbooks_scraper.py --limit 100             # scrape first 100 books
  python dcbooks_scraper.py --resume                # resume from saved progress
  python dcbooks_scraper.py --output my_output.xlsx # custom output filename
"""

import re, time, json, argparse, logging, os, sys
import html as html_mod
import base64, subprocess
from datetime import datetime
from pathlib import Path

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger(__name__)

SITEMAP_URL = "https://dcbookstore.com/sitemap.xml"
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) "
                  "Chrome/120.0.0.0 Safari/537.36",
}
DELAY = 0.5  # seconds between requests
PROGRESS_FILE = "dcbooks_progress.json"

# Wikidata mappings
LANG_QID = {
    "malayalam": "Q36236", "english": "Q1860", "hindi": "Q11051",
    "tamil": "Q5885", "kannada": "Q33673", "arabic": "Q13955",
    "telugu": "Q8097", "bengali": "Q9610", "marathi": "Q1571",
}

GENRE_QID = {
    "novel": "Q8261", "novels": "Q8261",
    "short stories": "Q49084", "short story": "Q49084", "story": "Q49084",
    "poetry": "Q482", "poems": "Q482",
    "biography": "Q36279", "autobiography": "Q4184",
    "autobiography biography": "Q36279",
    "autobiographies & biographies": "Q36279",
    "essays": "Q35760", "essay": "Q35760",
    "articles": "Q35760", "articles & jottings": "Q35760",
    "articles jottings": "Q35760",
    "study": "Q7075", "text book": "Q83790", "academics": "Q7075",
    "children's literature": "Q21077609", "children's books": "Q21077609",
    "childrens literature": "Q21077609", "childrens books": "Q21077609",
    "self help": "Q217467", "self-help": "Q217467",
    "history": "Q309", "philosophy": "Q5891", "religion": "Q9174",
    "travel & travelogues": "Q1075302", "travel": "Q1075302",
    "reference": "Q13136", "memoirs": "Q234460", "memoir": "Q234460",
    "crime thrillers": "Q19719828", "thriller": "Q19719828",
    "screenplays": "Q3328821", "screenplay": "Q3328821",
    "literary fiction": "Q1400336", "literary criticism & study": "Q58735",
    "popular science": "Q1093829", "science": "Q336",
    "photography": "Q125191", "art": "Q36649",
    "humor": "Q49076", "humour": "Q49076",
    "drama": "Q25379", "play": "Q25379",
    "career guidance": "Q18123741",
    "business management": "Q185451",
    "agriculture gardening": "Q11451",
    "architecture vasthu": "Q12271",
    "astrology": "Q34362", "astronomy": "Q333",
    "comic": "Q1004", "comics": "Q1004",
    "cooking": "Q185451", "cook book": "Q30070318",
    "dictionary": "Q23622", "encyclopedia": "Q5292",
    "health": "Q12147", "fitness": "Q309252",
    "law": "Q7748", "political science": "Q36442",
    "music": "Q638", "sports": "Q349",
    "women's writing": "Q15708736",
    "agriculture & gardening": "Q11451", "agriculture gardening": "Q11451",
    "autobiography & biography": "Q36279", "autobiography biography": "Q36279",
    "business & management": "Q185451", "business management": "Q185451",
    "cinema": "Q11424", "film": "Q11424",
    "collections & selected works": "Q20540385",
    "cookery": "Q30070318", "cooking": "Q30070318",
    "epics & myths": "Q8436", "epic": "Q8436",
    "health & fitness": "Q12147", "health fitness": "Q12147",
    "life style": "Q32090",
    "mathematics": "Q395", "psychology": "Q9418",
    "politics": "Q7163", "political science": "Q36442",
    "romance": "Q858330",
    "society & culture": "Q11042",
    "spirituality & mysticism": "Q131841",
    "translations": "Q7553",
    "travel & travelogue": "Q1075302", "travelogue": "Q1075302",
    "translation fest 2025": "", "best sellers": "",
    "vocal for local": "", "47th anniversary": "",
    "50 off": "", "50% off": "", "48 years": "",
    "anaswarakathakal": "", "book deals of the day": "",
    "books to buy under 100": "", "childrens 15 off": "",
    "childrens english literature festival": "",
    "karkidakam special": "", "rush hours": "",
    "previous edition books": "",
}


# ── Sucuri WAF bypass ──────────────────────────────────────────────────

def solve_sucuri(session):
    """Solve the Sucuri CloudProxy JavaScript challenge."""
    log.info("Solving Sucuri WAF challenge...")
    resp = session.get("https://dcbookstore.com/", headers=HEADERS, timeout=15)
    match = re.search(r"S='([^']+)'", resp.text)
    if not match:
        if resp.status_code == 200 and len(resp.text) > 5000:
            log.info("No Sucuri challenge detected — already accessible")
            return True
        log.warning("Could not find Sucuri challenge")
        return False

    decoded = base64.b64decode(match.group(1)).decode()
    js = (
        f"var document={{cookie:''}};var location={{reload:function(){{}}}};"
        f"{decoded};"
        f"console.log(JSON.stringify({{cookie:document.cookie}}));"
    )
    result = subprocess.run(["node", "-e", js], capture_output=True, text=True, timeout=10)
    if result.returncode != 0:
        log.error(f"Node.js error: {result.stderr}")
        return False

    data = json.loads(result.stdout)
    cookie_str = data["cookie"]
    name, val = cookie_str.split("=", 1)
    val = val.split(";")[0]
    session.cookies.set(name, val, domain="dcbookstore.com", path="/")
    log.info(f"Sucuri cookie set: {name}={val[:20]}...")
    return True


# ── Sitemap parsing ────────────────────────────────────────────────────

def get_book_urls_from_sitemap(session):
    """Parse sitemap.xml to get all book URLs."""
    log.info("Fetching sitemap.xml...")
    resp = session.get(SITEMAP_URL, headers=HEADERS, timeout=30)
    resp.raise_for_status()
    urls = re.findall(r"<loc>(https://dcbookstore\.com/books/[^<]+)</loc>", resp.text)
    log.info(f"Found {len(urls)} book URLs in sitemap")
    return urls


# ── Book page scraping ─────────────────────────────────────────────────

def is_malayalam(text):
    return bool(text) and any("\u0D00" <= c <= "\u0D7F" for c in text)


def scrape_book_page(url, session):
    """Scrape full metadata from a dcbookstore.com book page."""
    time.sleep(DELAY)
    try:
        resp = session.get(url, headers=HEADERS, timeout=20)
    except requests.RequestException as e:
        log.debug(f"Request failed: {url}: {e}")
        return None

    if resp.status_code != 200:
        return None

    text_content = resp.text
    if "sucuri_cloudproxy" in text_content.lower() and len(text_content) < 3000:
        # Need to re-solve the WAF
        if solve_sucuri(session):
            try:
                resp = session.get(url, headers=HEADERS, timeout=20)
                text_content = resp.text
            except:
                return None
        else:
            return None

    soup = BeautifulSoup(text_content, "lxml")
    text = soup.get_text(separator="\n")

    book = {"store_url": url}

    # Extract structured fields using regex on visible text
    field_map = {
        "title":    r"Book\s*:\s*(.+?)(?:\n|$)",
        "author":   r"Author:\s*(.+?)(?:\n|$)",
        "category": r"Category\s*:\s*(.+?)(?:\n|$)",
        "isbn":     r"ISBN\s*:\s*(\d[\d\-]+\d)(?:\n|$)",
        "binding":  r"Binding\s*:\s*(.+?)(?:\n|$)",
        "pub_date": r"Publishing Date\s*:\s*(.+?)(?:\n|$)",
        "publisher":r"Publisher\s*:\s*(.+?)(?:\n|$)",
        "edition":  r"Edition\s*:\s*(.+?)(?:\n|$)",
        "pages":    r"Number of pages\s*:\s*(\d+)",
        "language": r"Language\s*:\s*(.+?)(?:\n|$)",
    }
    for key, pattern in field_map.items():
        m = re.search(pattern, text, re.IGNORECASE)
        if m:
            book[key] = m.group(1).strip()

    # Separate ISBN-13 / ISBN-10
    isbn = book.pop("isbn", "")
    isbn_clean = isbn.replace("-", "")
    if len(isbn_clean) == 13:
        book["isbn13"] = isbn_clean
    elif len(isbn_clean) == 10:
        book["isbn10"] = isbn_clean
    elif isbn_clean:
        book["isbn13"] = isbn_clean

    # Extract Book Summary
    summary_match = re.search(r"Book Summary\s*\n(.+?)(?:\n\n|\nRELATED|\nNo Review|\Z)", text, re.DOTALL)
    if summary_match:
        book["summary"] = summary_match.group(1).strip()[:500]

    # Extract Malayalam text for title/description
    ml_chunks = re.findall(r"[\u0D00-\u0D7F][\u0D00-\u0D7F\s\u200C\u200D.,;!?]+", text)
    if ml_chunks:
        # The first substantial chunk is often a Malayalam description
        for chunk in ml_chunks:
            if len(chunk.strip()) > 5:
                book["summary_ml"] = chunk.strip()[:200]
                break

    # Cover image
    og_img = soup.select_one('meta[property="og:image"]')
    if og_img:
        book["cover_url"] = og_img.get("content", "")
    else:
        img = soup.select_one(".productImage img, .book-image img, img[src*='uploads']")
        if img:
            book["cover_url"] = img.get("src", "")

    return book if book.get("title") else None


# ── Wikidata formatting ───────────────────────────────────────────────

def normalize_language(lang_str):
    if not lang_str:
        return ""
    return LANG_QID.get(lang_str.strip().lower(), lang_str)

def normalize_genre(cat_str):
    if not cat_str:
        return ""
    cats = [c.strip() for c in re.split(r"[,;/]", cat_str)]
    results = []
    for cat in cats:
        q = GENRE_QID.get(cat.lower().strip(), None)
        if q is None:
            results.append(cat.strip())
        elif q:
            results.append(q)
        # skip empty string mappings (promotional categories)
    return " | ".join(r for r in results if r)

def normalize_date(date_str):
    if not date_str:
        return ""
    for fmt in ("%d-%m-%Y", "%d-%m-%y", "%Y-%m-%d", "%d/%m/%Y", "%m-%d-%Y"):
        try:
            dt = datetime.strptime(date_str.strip(), fmt)
            return f"+{dt.strftime('%Y-%m-%d')}T00:00:00Z/11"
        except ValueError:
            continue
    m = re.search(r"(\d{4})", date_str)
    if m:
        return f"+{m.group(1)}-00-00T00:00:00Z/9"
    return date_str

def prepare_wikidata_row(book):
    title = book.get("title", "")
    lang = book.get("language", "")
    row = {
        "qid": "",
        "Label (en)": title,
        "Label (ml)": "",
        "Description (en)": "",
        "Description (ml)": book.get("summary_ml", ""),
        "P31 (instance of)": "Q47461344",
        "P50 (author)": book.get("author", ""),
        "P212 (ISBN-13)": book.get("isbn13", ""),
        "P957 (ISBN-10)": book.get("isbn10", ""),
        "P123 (publisher)": book.get("publisher", ""),
        "P407 (language of work)": normalize_language(lang),
        "P495 (country of origin)": "Q668",
        "P577 (publication date)": normalize_date(book.get("pub_date", "")),
        "P1104 (number of pages)": book.get("pages", ""),
        "P136 (genre)": normalize_genre(book.get("category", "")),
        "P393 (edition number)": book.get("edition", ""),
        "P437 (distribution format)": book.get("binding", ""),
        "S248 (stated in)": "DC Books website",
        "S854 (reference URL)": book.get("store_url", ""),
        "cover_image_url": book.get("cover_url", ""),
    }
    # Auto-description
    parts = []
    if lang:
        parts.append(f"{lang}-language")
    cat = book.get("category", "")
    if cat:
        first_cat = [c.strip().lower() for c in cat.split(",") if c.strip().lower() not in
                     ("best sellers", "vocal for local", "47th anniversary", "48 years",
                      "50 off", "translation fest 2025", "book deals of the day",
                      "books to buy under 100", "childrens 15 off",
                      "childrens english literature festival", "anaswarakathakal")]
        if first_cat:
            parts.append(first_cat[0])
        else:
            parts.append("book")
    else:
        parts.append("book")
    pub = book.get("publisher", "")
    if pub:
        parts.append(f"published by {pub}")
    row["Description (en)"] = " ".join(parts)
    return row


# ── Excel output ──────────────────────────────────────────────────────

COLUMNS = [
    "qid", "Label (en)", "Label (ml)", "Description (en)", "Description (ml)",
    "P31 (instance of)", "P50 (author)", "P212 (ISBN-13)", "P957 (ISBN-10)",
    "P123 (publisher)", "P407 (language of work)", "P495 (country of origin)",
    "P577 (publication date)", "P1104 (number of pages)", "P136 (genre)",
    "P393 (edition number)", "P437 (distribution format)",
    "S248 (stated in)", "S854 (reference URL)", "cover_image_url",
]

RAW_COLS = [
    "title", "author", "isbn13", "isbn10", "publisher", "language",
    "category", "binding", "pub_date", "edition", "pages",
    "summary", "summary_ml", "cover_url", "store_url",
]

LEGEND = [
    ("Column", "Wikidata Property", "Description", "Example"),
    ("qid", "-", "Leave blank for new items; use Qxxx to update existing", "Q12345678"),
    ("Label (en)", "Label", "English/transliterated title", "Aadujeevitham"),
    ("Label (ml)", "Label", "Malayalam title (add manually if known)", "ആടുജീവിതം"),
    ("Description (en)", "Description", "Short English description (auto-generated)", "Malayalam-language novel published by DC Books"),
    ("Description (ml)", "Description", "Malayalam description (from book summary)", "മലയാള നോവൽ..."),
    ("P31", "instance of", "Q47461344 = written work", "Q47461344"),
    ("P50", "author", "Author name (replace with QID if Wikidata item exists)", "Benyamin"),
    ("P212", "ISBN-13", "13-digit ISBN", "9788126435746"),
    ("P957", "ISBN-10", "10-digit ISBN if available", "8126435747"),
    ("P123", "publisher", "Publisher name (Q5203520 = DC Books on Wikidata)", "DC Books"),
    ("P407", "language of work", "Q36236=Malayalam, Q1860=English", "Q36236"),
    ("P495", "country of origin", "Q668 = India", "Q668"),
    ("P577", "publication date", "+YYYY-MM-DDT00:00:00Z/precision", "+2012-01-01T00:00:00Z/11"),
    ("P1104", "number of pages", "Integer page count", "224"),
    ("P136", "genre", "Q8261=novel, Q482=poetry, Q49084=short stories", "Q8261"),
    ("P393", "edition number", "Edition number or label", "3"),
    ("P437", "distribution format", "Physical format", "Paperback"),
    ("S248", "stated in", "Reference source", "DC Books website"),
    ("S854", "reference URL", "URL to the source page", "https://dcbookstore.com/books/..."),
    ("cover_image_url", "-", "Cover image URL (for Wikimedia Commons)", "https://dcbookstore.com/uploads/..."),
]


def write_excel(books, output_path):
    wb = Workbook()
    thin = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
    hf = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    df = Font(name="Arial", size=10)
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    da = Alignment(vertical="center", wrap_text=True)

    # Sheet 1: Wikidata Upload
    ws = wb.active
    ws.title = "Wikidata Upload"
    f1 = PatternFill("solid", fgColor="2B579A")
    for ci, col in enumerate(COLUMNS, 1):
        c = ws.cell(row=1, column=ci, value=col)
        c.font, c.fill, c.alignment, c.border = hf, f1, ha, thin
    rows = [prepare_wikidata_row(b) for b in books]
    for ri, rd in enumerate(rows, 2):
        for ci, col in enumerate(COLUMNS, 1):
            c = ws.cell(row=ri, column=ci, value=rd.get(col, ""))
            c.font, c.border, c.alignment = df, thin, da
    ws.freeze_panes = "A2"
    last_col = chr(64 + len(COLUMNS)) if len(COLUMNS) <= 26 else "T"
    ws.auto_filter.ref = f"A1:{last_col}{len(rows)+1}"
    widths = [8,35,35,45,40,18,25,18,15,25,18,18,22,12,25,10,15,18,45,45]
    for i, w in enumerate(widths[:len(COLUMNS)]):
        col_letter = chr(65+i) if i < 26 else chr(64+i//26) + chr(65+i%26)
        ws.column_dimensions[col_letter].width = w

    # Sheet 2: Raw Data
    ws2 = wb.create_sheet("Raw Data")
    f2 = PatternFill("solid", fgColor="4A7C59")
    for ci, col in enumerate(RAW_COLS, 1):
        c = ws2.cell(row=1, column=ci, value=col)
        c.font, c.fill, c.alignment, c.border = hf, f2, ha, thin
    for ri, book in enumerate(books, 2):
        for ci, col in enumerate(RAW_COLS, 1):
            c = ws2.cell(row=ri, column=ci, value=book.get(col, ""))
            c.font, c.border = df, thin
    ws2.freeze_panes = "A2"
    raw_widths = [35,25,18,15,25,12,30,12,15,10,8,50,50,45,45]
    for i, w in enumerate(raw_widths[:len(RAW_COLS)]):
        ws2.column_dimensions[chr(65+i)].width = w

    # Sheet 3: Legend
    ws3 = wb.create_sheet("Wikidata Property Legend")
    f3 = PatternFill("solid", fgColor="8B4513")
    for ri, rd in enumerate(LEGEND, 1):
        for ci, val in enumerate(rd, 1):
            c = ws3.cell(row=ri, column=ci, value=val)
            if ri == 1:
                c.font, c.fill, c.alignment = hf, f3, ha
            else:
                c.font = df
            c.border = thin
    for i, w in enumerate([18, 22, 55, 40], 1):
        ws3.column_dimensions[chr(64+i)].width = w

    wb.save(output_path)
    log.info(f"Saved {len(books)} books to {output_path}")


# ── Progress save/resume ──────────────────────────────────────────────

def save_progress(books, scraped_urls, progress_file=PROGRESS_FILE):
    with open(progress_file, "w", encoding="utf-8") as f:
        json.dump({"books": books, "scraped_urls": list(scraped_urls)}, f, ensure_ascii=False)

def load_progress(progress_file=PROGRESS_FILE):
    if not os.path.exists(progress_file):
        return [], set()
    with open(progress_file, "r", encoding="utf-8") as f:
        data = json.load(f)
    return data.get("books", []), set(data.get("scraped_urls", []))


# ── Main ──────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Scrape DC Books → Wikidata Excel")
    parser.add_argument("--limit", type=int, default=0,
                        help="Max books to scrape (0 = all)")
    parser.add_argument("--resume", action="store_true",
                        help="Resume from saved progress")
    parser.add_argument("--output", default="dcbooks_wikidata.xlsx")
    parser.add_argument("--save-every", type=int, default=50,
                        help="Save progress every N books")
    args = parser.parse_args()

    session = requests.Session()

    # Step 1: Get all book URLs from sitemap
    book_urls = get_book_urls_from_sitemap(session)

    if args.limit > 0:
        book_urls = book_urls[:args.limit]
        log.info(f"Limiting to {args.limit} books")

    # Step 2: Solve Sucuri WAF
    if not solve_sucuri(session):
        log.error("Failed to solve Sucuri WAF. Cannot proceed.")
        sys.exit(1)

    # Step 3: Resume or start fresh
    if args.resume:
        all_books, scraped_urls = load_progress()
        log.info(f"Resumed: {len(all_books)} books already scraped")
    else:
        all_books, scraped_urls = [], set()

    # Step 4: Scrape each book page
    remaining = [u for u in book_urls if u not in scraped_urls]
    log.info(f"Scraping {len(remaining)} book pages...")

    failed = 0
    for i, url in enumerate(remaining):
        if (i + 1) % 20 == 0 or i == 0:
            log.info(f"  Progress: {i+1}/{len(remaining)} "
                     f"(total scraped: {len(all_books)}, failed: {failed})")

        book = scrape_book_page(url, session)
        scraped_urls.add(url)
        if book:
            all_books.append(book)
        else:
            failed += 1

        # Save progress periodically
        if (i + 1) % args.save_every == 0:
            save_progress(all_books, scraped_urls)
            log.info(f"  Progress saved ({len(all_books)} books)")

        # Re-solve WAF every 200 requests just in case
        if (i + 1) % 200 == 0:
            solve_sucuri(session)

    # Step 5: Write Excel
    save_progress(all_books, scraped_urls)
    write_excel(all_books, args.output)

    # Summary
    log.info(f"\n{'='*50}")
    log.info(f"DONE! Total books scraped: {len(all_books)}")
    log.info(f"Failed pages: {failed}")
    langs = {}
    for b in all_books:
        l = b.get("language", "Unknown")
        langs[l] = langs.get(l, 0) + 1
    log.info(f"Languages: {langs}")
    with_isbn = sum(1 for b in all_books if b.get("isbn13") or b.get("isbn10"))
    with_summary = sum(1 for b in all_books if b.get("summary"))
    with_ml = sum(1 for b in all_books if b.get("summary_ml"))
    log.info(f"With ISBN: {with_isbn}/{len(all_books)}")
    log.info(f"With summary: {with_summary}, With Malayalam text: {with_ml}")


if __name__ == "__main__":
    main()
