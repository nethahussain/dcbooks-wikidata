#!/usr/bin/env python3
"""
Kerala Book Store Scraper — DC Books Malayalam titles
=====================================================
Scrapes keralabookstore.com for DC Books titles to get:
- Malayalam script book titles
- Malayalam script author names
- Malayalam category names
- ISBN
- Book ID and URL

Then merges with the existing dcbooks_wikidata.xlsx to fill in
Malayalam titles for books matched by ISBN.

keralabookstore.com blocks cloud/datacenter IPs — run this locally.

Usage:
  python3 scrape_keralabookstore.py
  python3 scrape_keralabookstore.py --merge data/dcbooks_wikidata.xlsx
"""

import requests, re, time, json, argparse, os, sys
from bs4 import BeautifulSoup
from urllib.parse import urlencode

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) "
                  "Chrome/120.0.0.0 Safari/537.36",
    "Accept-Language": "ml,en;q=0.9",
}
BASE = "https://keralabookstore.com"
SEARCH_URL = f"{BASE}/saved-search.do"
DELAY = 1.5
PROGRESS_FILE = "kbs_progress.json"
OUTPUT_FILE = "keralabookstore_dcbooks.json"


def get_soup(url, session):
    time.sleep(DELAY)
    resp = session.get(url, headers=HEADERS, timeout=60)
    resp.raise_for_status()
    return BeautifulSoup(resp.text, "lxml"), resp.text


def save_progress(books, page, scraped_ids):
    with open(PROGRESS_FILE, "w", encoding="utf-8") as f:
        json.dump({"books": books, "page": page, "scraped_ids": list(scraped_ids)}, f, ensure_ascii=False)


def load_progress():
    if os.path.exists(PROGRESS_FILE):
        with open(PROGRESS_FILE, encoding="utf-8") as f:
            data = json.load(f)
        return data.get("books", []), data.get("page", 1), set(data.get("scraped_ids", []))
    return [], 1, set()


def extract_book_from_listing(item):
    """Extract book metadata from a listing item on the search results page."""
    book = {}

    # Title (Malayalam) — usually in an <a> or heading
    title_el = item.select_one("a.bookTitle, .book-title a, h3 a, h4 a, .title a, a[href*='/book/']")
    if not title_el:
        # Fallback: find any link to a book page
        for a in item.select("a[href]"):
            href = a.get("href", "")
            if "/book/" in href:
                title_el = a
                break

    if title_el:
        book["title_ml"] = title_el.get_text(strip=True)
        href = title_el.get("href", "")
        if href:
            if href.startswith("/"):
                href = BASE + href
            book["url"] = href
            # Extract book ID from URL pattern /book/{slug}/{id}/
            id_match = re.search(r"/book/[^/]+/(\d+)", href)
            if id_match:
                book["book_id"] = id_match.group(1)

    # Author (Malayalam)
    author_el = item.select_one(".author, .book-author, a[href*='/books-by/']")
    if author_el:
        book["author_ml"] = author_el.get_text(strip=True)

    # ISBN
    isbn_el = item.select_one(".isbn, .book-isbn")
    if isbn_el:
        isbn_text = isbn_el.get_text(strip=True)
        isbn_match = re.search(r"(\d{10,13})", isbn_text)
        if isbn_match:
            book["isbn"] = isbn_match.group(1)

    # Price
    price_el = item.select_one(".price, .book-price, .amount")
    if price_el:
        book["price"] = price_el.get_text(strip=True)

    # Publisher
    pub_el = item.select_one(".publisher, a[href*='publisher=']")
    if pub_el:
        book["publisher"] = pub_el.get_text(strip=True)

    # If we couldn't find structured elements, try parsing text
    text = item.get_text(separator="|", strip=True)
    if "isbn" not in book:
        isbn_m = re.search(r"ISBN[:\s]*(\d{10,13})", text, re.IGNORECASE)
        if isbn_m:
            book["isbn"] = isbn_m.group(1)

    return book if book.get("title_ml") or book.get("url") else None


def scrape_book_page(url, session):
    """Scrape a single book page for detailed metadata."""
    try:
        soup, text = get_soup(url, session)
    except Exception as e:
        print(f"  Error: {e}")
        return None

    book = {"url": url}

    # The <title> tag has rich metadata:
    # "buy the book {ML_TITLE} written by {ML_AUTHOR} in category {ML_CAT}, ISBN {ISBN}, Published by {PUB}"
    title_tag = soup.title.string if soup.title else ""
    if title_tag:
        # Malayalam title from <title>
        m = re.search(r"buy the book (.+?) written by", title_tag)
        if m:
            book["title_ml"] = m.group(1).strip()

        # Malayalam author
        m = re.search(r"written by (.+?) in category", title_tag)
        if m:
            book["author_ml"] = m.group(1).strip()

        # Category in Malayalam
        m = re.search(r"in category (.+?),", title_tag)
        if m:
            book["category_ml"] = m.group(1).strip()

        # ISBN
        m = re.search(r"ISBN (\d{10,13})", title_tag)
        if m:
            book["isbn"] = m.group(1)

        # Publisher
        m = re.search(r"Published by (.+?)(?:\s+from|\s*$)", title_tag)
        if m:
            book["publisher"] = m.group(1).strip()

    # Also look for structured data in the page body
    page_text = soup.get_text(separator="\n")

    # Find English title (if present)
    for pattern in [r"English Title[:\s]*(.+?)(?:\n|$)", r"Original Title[:\s]*(.+?)(?:\n|$)"]:
        m = re.search(pattern, page_text, re.IGNORECASE)
        if m:
            book["title_en"] = m.group(1).strip()

    # Find ISBN if not from title
    if "isbn" not in book:
        m = re.search(r"ISBN[:\s]*(\d{10,13})", page_text, re.IGNORECASE)
        if m:
            book["isbn"] = m.group(1)

    # Find all Malayalam text blocks for description
    ml_chunks = re.findall(r"[\u0D00-\u0D7F][\u0D00-\u0D7F\s\u200C\u200D.,;!?()]+", page_text)
    if ml_chunks:
        # The longest chunk is usually the description
        longest = max(ml_chunks, key=len)
        if len(longest) > 20:
            book["description_ml"] = longest.strip()[:500]

    # Book ID from URL
    id_match = re.search(r"/book/[^/]+/(\d+)", url)
    if id_match:
        book["book_id"] = id_match.group(1)

    return book if (book.get("title_ml") or book.get("isbn")) else None


def scrape_search_results(session):
    """Scrape the publisher search results page for DC Books."""
    all_books = []
    seen_ids = set()
    page = 1

    print("Scraping DC Books listings from keralabookstore.com...")

    # First, try the publisher search page
    url = f"{SEARCH_URL}?publisher=DC+Books"
    print(f"  Fetching: {url}")

    try:
        soup, text = get_soup(url, session)
    except Exception as e:
        print(f"  Failed to load search page: {e}")
        print("  Tip: keralabookstore.com may block datacenter IPs. Run this from your home network.")
        return all_books

    # Find all book links on the page
    book_links = []
    for a in soup.select("a[href*='/book/']"):
        href = a.get("href", "")
        if "/book/" in href and href not in book_links:
            if href.startswith("/"):
                href = BASE + href
            book_links.append(href)

    print(f"  Found {len(book_links)} book links on search page")

    # Check for pagination
    page_links = soup.select("a[href*='page='], a[href*='pageNo='], .pagination a, .page-link")
    max_page = 1
    for pl in page_links:
        href = pl.get("href", "")
        text_val = pl.get_text(strip=True)
        for pattern in [r"page=(\d+)", r"pageNo=(\d+)", r"start=(\d+)"]:
            m = re.search(pattern, href)
            if m:
                max_page = max(max_page, int(m.group(1)))
        if text_val.isdigit():
            max_page = max(max_page, int(text_val))

    # Also check for "Next" links
    next_link = soup.select_one("a.next, a:contains('Next'), a:contains('>>'), a[rel='next']")

    print(f"  Detected {max_page} pages of results")
    if next_link:
        print(f"  Next link found: {next_link.get('href', '')}")

    # If pagination exists, scrape remaining pages
    if max_page > 1 or next_link:
        for pg in range(2, max_page + 1):
            # Try common pagination patterns
            for param in [f"page={pg}", f"pageNo={pg}", f"start={(pg-1)*20}"]:
                pg_url = f"{SEARCH_URL}?publisher=DC+Books&{param}"
                try:
                    soup2, _ = get_soup(pg_url, session)
                    new_links = []
                    for a in soup2.select("a[href*='/book/']"):
                        href = a.get("href", "")
                        if "/book/" in href:
                            if href.startswith("/"):
                                href = BASE + href
                            if href not in book_links:
                                new_links.append(href)
                                book_links.append(href)
                    if new_links:
                        print(f"  Page {pg}: +{len(new_links)} links (total: {len(book_links)})")
                        break  # This pagination pattern worked
                except:
                    continue

    # Now scrape each book page
    print(f"\nScraping {len(book_links)} individual book pages...")
    for i, link in enumerate(book_links):
        id_match = re.search(r"/book/[^/]+/(\d+)", link)
        book_id = id_match.group(1) if id_match else link
        if book_id in seen_ids:
            continue
        seen_ids.add(book_id)

        book = scrape_book_page(link, session)
        if book:
            all_books.append(book)

        if (i + 1) % 50 == 0:
            print(f"  {i+1}/{len(book_links)} scraped, {len(all_books)} books collected")
            save_progress(all_books, page, seen_ids)

    return all_books


def scrape_by_category(session):
    """Alternative: scrape books by browsing categories."""
    all_books = []
    seen_ids = set()

    # Common category URLs for DC Books
    categories_to_try = [
        "/books/category/%E0%B4%A8%E0%B5%8B%E0%B4%B5%E0%B5%BD/1/",  # Novel
        "/books/category/%E0%B4%95%E0%B4%B5%E0%B4%BF%E0%B4%A4/2/",  # Poetry
        "/books/category/%E0%B4%9A%E0%B5%86%E0%B4%B1%E0%B5%81%E0%B4%95%E0%B4%A5/40/",  # Short stories
    ]

    for cat_path in categories_to_try:
        url = BASE + cat_path
        try:
            soup, _ = get_soup(url, session)
            links = [a.get("href", "") for a in soup.select("a[href*='/book/']")]
            print(f"Category {cat_path}: {len(links)} links")
        except:
            pass

    return all_books


def merge_with_excel(books, excel_path):
    """Merge scraped Malayalam data into the existing Excel file by ISBN."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("Install openpyxl: pip install openpyxl")
        return

    # Build ISBN lookup from scraped data
    isbn_map = {}
    for b in books:
        isbn = b.get("isbn", "")
        if isbn:
            isbn_map[isbn] = b

    print(f"\nMerging {len(isbn_map)} ISBN-matched books into {excel_path}...")

    wb = load_workbook(excel_path)
    ws = wb["Wikidata Upload"]

    # Find columns
    headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    col_isbn13 = headers.get("P212 (ISBN-13)")
    col_isbn10 = headers.get("P957 (ISBN-10)")
    col_label_ml = headers.get("Label (ml)")
    col_desc_ml = headers.get("Description (ml)")

    # Add author Malayalam column if not present
    col_author_ml = headers.get("P50_ml (author Malayalam)")
    if not col_author_ml:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        col_author_ml = ws.max_column + 1
        cell = ws.cell(1, col_author_ml, value="P50_ml (author Malayalam)")
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        cell.fill = PatternFill("solid", fgColor="2B579A")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=Side("thin"), right=Side("thin"),
                            top=Side("thin"), bottom=Side("thin"))

    matched = 0
    ml_updated = 0
    for r in range(2, ws.max_row + 1):
        isbn13 = str(ws.cell(r, col_isbn13).value or "").strip()
        isbn10 = str(ws.cell(r, col_isbn10).value or "").strip()

        kbs_book = isbn_map.get(isbn13) or isbn_map.get(isbn10)
        if not kbs_book:
            continue

        matched += 1

        # Update Malayalam title if keralabookstore has a real Malayalam one
        ml_title = kbs_book.get("title_ml", "")
        if ml_title and any("\u0D00" <= c <= "\u0D7F" for c in ml_title):
            existing = ws.cell(r, col_label_ml).value or ""
            # Prefer the keralabookstore version (actual Malayalam) over transliteration
            ws.cell(r, col_label_ml).value = ml_title
            ml_updated += 1

        # Update Malayalam author
        ml_author = kbs_book.get("author_ml", "")
        if ml_author and any("\u0D00" <= c <= "\u0D7F" for c in ml_author):
            ws.cell(r, col_author_ml).value = ml_author

        # Update Malayalam description if we have a better one
        ml_desc = kbs_book.get("description_ml", "")
        if ml_desc and len(ml_desc) > 20 and col_desc_ml:
            existing_desc = ws.cell(r, col_desc_ml).value or ""
            if not existing_desc or len(ml_desc) > len(existing_desc):
                ws.cell(r, col_desc_ml).value = ml_desc

    wb.save(excel_path)
    print(f"Matched: {matched}/{ws.max_row - 1} books by ISBN")
    print(f"Malayalam titles updated: {ml_updated}")
    print(f"Saved to {excel_path}")


def main():
    parser = argparse.ArgumentParser(description="Scrape keralabookstore.com DC Books")
    parser.add_argument("--merge", type=str, default="",
                        help="Path to dcbooks_wikidata.xlsx to merge Malayalam titles into")
    parser.add_argument("--resume", action="store_true",
                        help="Resume from saved progress")
    args = parser.parse_args()

    session = requests.Session()

    if args.resume:
        books, page, scraped_ids = load_progress()
        print(f"Resumed: {len(books)} books from progress")
    else:
        books = []

    if not books:
        books = scrape_search_results(session)

    # Save results
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(books, f, ensure_ascii=False, indent=1)
    print(f"\nSaved {len(books)} books to {OUTPUT_FILE}")

    # Summary
    with_ml_title = sum(1 for b in books if b.get("title_ml") and
                        any("\u0D00" <= c <= "\u0D7F" for c in b.get("title_ml", "")))
    with_isbn = sum(1 for b in books if b.get("isbn"))
    with_author_ml = sum(1 for b in books if b.get("author_ml") and
                         any("\u0D00" <= c <= "\u0D7F" for c in b.get("author_ml", "")))
    print(f"With Malayalam title: {with_ml_title}")
    print(f"With ISBN: {with_isbn}")
    print(f"With Malayalam author: {with_author_ml}")

    # Merge if requested
    if args.merge and os.path.exists(args.merge):
        merge_with_excel(books, args.merge)
    elif args.merge:
        print(f"File not found: {args.merge}")

    # Show samples
    print("\nSample entries:")
    for b in books[:10]:
        ml = b.get("title_ml", "")
        author = b.get("author_ml", "")
        isbn = b.get("isbn", "")
        print(f"  {ml[:40]:40s} | {author[:25]:25s} | {isbn}")


if __name__ == "__main__":
    main()
